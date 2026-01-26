from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
from openpyxl import load_workbook
from io import BytesIO
import re
from email_validator import validate_email, EmailNotValidError
from app import schemas
from app.database import get_db
from app.models import Student

router = APIRouter()


@router.post("/students")
async def import_students_from_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be Excel format (.xlsx or .xls)")
    
    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active
        
        def normalize_header(value: object) -> str:
            if value is None:
                return ""
            return re.sub(r"\s+", " ", str(value).strip().lower())

        def title_case(value: str) -> str:
            parts = re.sub(r"\s+", " ", value.strip()).split(" ")
            return " ".join([p.capitalize() for p in parts if p])

        def normalize_phone(value: str) -> tuple[str, str]:
            raw = value.strip()
            cleaned = re.sub(r"[^\d+]", "", raw)
            digits = re.sub(r"\D", "", cleaned)
            if len(digits) < 10 or len(digits) > 15:
                raise ValueError("некорректный номер телефона")
            normalized = f"+{digits}" if raw.startswith("+") else digits
            return normalized, digits

        header_rows = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if not header_rows:
            raise HTTPException(status_code=400, detail="Файл пустой или без заголовков")
        header_row = header_rows[0]
        expected_headers = [
            "фио школьника",
            "номер телефона школьника",
            "email школьника",
            "название школы",
            "фио родителя",
            "номер телефона родителя",
            "название курса",
        ]
        for idx, expected in enumerate(expected_headers[:6]):
            actual = normalize_header(header_row[idx]) if idx < len(header_row) else ""
            if actual != expected:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Неверный заголовок в колонке {idx + 1}. "
                        f"Ожидалось: \"{expected_headers[idx]}\", "
                        f"получено: \"{header_row[idx] if idx < len(header_row) else ''}\""
                    ),
                )

        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        imported_students = []
        errors = []

        existing_emails = {
            (email or "").strip().lower()
            for (email,) in db.query(Student.email).all()
            if email
        }
        existing_phones = {
            re.sub(r"\D", "", (phone or "")) for (phone,) in db.query(Student.phone).all() if phone
        }
        existing_parent_phones = {
            re.sub(r"\D", "", (phone or "")) for (phone,) in db.query(Student.parent_phone).all() if phone
        }
        seen_emails = set()
        seen_phones = set()
        seen_parent_phones = set()
        
        for idx, row in enumerate(rows, start=2):
            if not any(row):
                continue
            
            try:
                if len(row) < 6:
                    errors.append(f"Строка {idx}: недостаточно данных")
                    continue
                
                full_name = str(row[0]).strip() if row[0] else None
                phone = str(row[1]).strip() if row[1] else None
                email = str(row[2]).strip() if row[2] else None
                school_name = str(row[3]).strip() if row[3] else None
                parent_name = str(row[4]).strip() if row[4] else None
                parent_phone = str(row[5]).strip() if row[5] else None
                desired_course_name = str(row[6]).strip() if len(row) > 6 and row[6] else None
                
                if not all([full_name, phone, email, school_name, parent_name, parent_phone]):
                    errors.append(f"Строка {idx}: отсутствуют обязательные поля")
                    continue

                full_name = title_case(full_name)
                school_name = title_case(school_name)
                parent_name = title_case(parent_name)
                email = email.strip().lower()
                if desired_course_name:
                    desired_course_name = title_case(desired_course_name)

                try:
                    validate_email(email, check_deliverability=False)
                except EmailNotValidError:
                    errors.append(f"Строка {idx}: некорректный email")
                    continue

                try:
                    phone_normalized, phone_digits = normalize_phone(phone)
                    parent_phone_normalized, parent_phone_digits = normalize_phone(parent_phone)
                except ValueError as exc:
                    errors.append(f"Строка {idx}: {str(exc)}")
                    continue

                if email in seen_emails or email in existing_emails:
                    errors.append(f"Строка {idx}: дубликат email")
                    continue
                if phone_digits in seen_phones or phone_digits in existing_phones:
                    errors.append(f"Строка {idx}: дубликат телефона школьника")
                    continue
                if parent_phone_digits in seen_parent_phones or parent_phone_digits in existing_parent_phones:
                    errors.append(f"Строка {idx}: дубликат телефона родителя")
                    continue

                seen_emails.add(email)
                seen_phones.add(phone_digits)
                seen_parent_phones.add(parent_phone_digits)
                
                student_data = {
                    "full_name": full_name,
                    "phone": phone_normalized,
                    "email": email,
                    "school_name": school_name,
                    "parent_name": parent_name,
                    "parent_phone": parent_phone_normalized,
                    "desired_course_name": desired_course_name
                }
                
                db_student = Student(**student_data)
                db.add(db_student)
                imported_students.append(db_student)
                
            except Exception as e:
                errors.append(f"Строка {idx}: ошибка обработки - {str(e)}")
        
        if imported_students:
            db.commit()
            for student in imported_students:
                db.refresh(student)
        
        if errors:
            return {
                "imported": [schemas.Student.model_validate(s) for s in imported_students],
                "errors": errors,
                "message": f"Импортировано {len(imported_students)} школьников. Ошибок: {len(errors)}"
            }
        
        return [schemas.Student.model_validate(s) for s in imported_students]
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка при обработке файла: {str(e)}")
